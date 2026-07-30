from flask import Blueprint, render_template, request, jsonify, session, flash, redirect, url_for, send_file
from flask_babel import gettext, get_locale
from utils.db import get_db_connection
from utils.auth import login_required
from utils.rbac import require_permission
from datetime import datetime

payroll_bp = Blueprint('payroll', __name__)


def _uid():
    return session.get('user_id')


# ================= Monthly Transactions =================
@payroll_bp.route('/payroll/transactions')
@login_required
@require_permission('salary.calculate')
def transactions():
    conn = get_db_connection()
    now = datetime.now()
    month = int(request.args.get('month', now.month))
    year = int(request.args.get('year', now.year))
    rows = conn.execute('''
        SELECT t.*, e.name AS employee_name, e.arabic_name, e.employee_number,
               it.code AS item_code, it.name_ar AS item_name_ar, it.name_en AS item_name_en, it.kind
        FROM payroll_transactions t
        JOIN employees e ON e.id = t.employee_id
        JOIN payroll_item_types it ON it.id = t.item_type_id
        WHERE t.month = ? AND t.year = ?
        ORDER BY t.created_at DESC
    ''', (month, year)).fetchall()
    item_types = conn.execute('''
        SELECT * FROM payroll_item_types
        WHERE species = 'transaction' AND is_active = 1 AND affects = 'monthly' AND is_system = 0
        ORDER BY kind, sort_order
    ''').fetchall()
    employees = conn.execute(
        "SELECT id, name, arabic_name, employee_number FROM employees WHERE is_active = 1 "
        "ORDER BY CAST(employee_number AS INTEGER), employee_number").fetchall()
    totals = {
        'earning': sum(r['amount'] for r in rows if r['kind'] == 'earning' and r['status'] == 'active'),
        'deduction': sum(r['amount'] for r in rows if r['kind'] == 'deduction' and r['status'] == 'active'),
    }
    return render_template('payroll_transactions.html', rows=rows, item_types=item_types,
                           employees=employees, month=month, year=year, totals=totals)


@payroll_bp.route('/payroll/transactions/add', methods=['POST'])
@login_required
@require_permission('salary.calculate')
def add_transaction():
    try:
        employee_id = int(request.form['employee_id'])
        item_type_id = int(request.form['item_type_id'])
        month = int(request.form['month'])
        year = int(request.form['year'])
        from utils.payroll_engine import is_period_locked
        if is_period_locked(get_db_connection(), month, year):
            flash(gettext('x.err_period_closed'), 'error')
            return redirect(url_for('payroll.transactions', month=month, year=year))
        amount = float(request.form['amount'])
        reason = request.form.get('reason', '').strip()
        if amount <= 0:
            flash(gettext('x.f_amount_positive'), 'error')
            return redirect(url_for('payroll.transactions', month=month, year=year))
        it = get_db_connection().execute(
            "SELECT id FROM payroll_item_types WHERE id = ? AND species = 'transaction' AND is_active = 1",
            (item_type_id,)).fetchone()
        if not it:
            flash(gettext('x.f_item_type_invalid'), 'error')
            return redirect(url_for('payroll.transactions', month=month, year=year))
        conn = get_db_connection()
        conn.execute('''
            INSERT INTO payroll_transactions (employee_id, item_type_id, month, year, amount, reason, entered_by)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (employee_id, item_type_id, month, year, amount, reason, _uid()))
        conn.commit()
        flash(gettext('x.f_transaction_added'), 'success')
    except Exception as e:
        flash(gettext('x.f_transaction_error') % {'p0': f'{str(e)}'}, 'error')
    return redirect(url_for('payroll.transactions', month=request.form.get('month'), year=request.form.get('year')))


@payroll_bp.route('/payroll/transactions/cancel/<int:id>', methods=['POST'])
@login_required
@require_permission('salary.calculate')
def cancel_transaction(id):
    conn = get_db_connection()
    row = conn.execute('SELECT month, year, status FROM payroll_transactions WHERE id = ?', (id,)).fetchone()
    if not row:
        flash(gettext('x.f_not_found'), 'error')
        return redirect(url_for('payroll.transactions'))
    from utils.payroll_engine import is_period_locked
    if is_period_locked(conn, row['month'], row['year']):
        flash(gettext('x.err_period_closed'), 'error')
        return redirect(url_for('payroll.transactions', month=row['month'], year=row['year']))
    if row['status'] == 'active':
        conn.execute('''UPDATE payroll_transactions
                        SET status = 'cancelled', cancelled_by = ?, cancelled_at = CURRENT_TIMESTAMP
                        WHERE id = ?''', (_uid(), id))
        conn.commit()
        flash(gettext('x.f_transaction_cancelled'), 'success')
    return redirect(url_for('payroll.transactions', month=row['month'], year=row['year']))


# ================= Loans =================
@payroll_bp.route('/payroll/loans')
@login_required
@require_permission('salary.calculate')
def loans():
    conn = get_db_connection()
    rows = conn.execute('''
        SELECT l.*, e.name AS employee_name, e.arabic_name, e.employee_number,
               COALESCE((SELECT SUM(amount) FROM loan_payments p WHERE p.loan_id = l.id), 0) AS paid
        FROM employee_loans l
        JOIN employees e ON e.id = l.employee_id
        ORDER BY CASE l.status WHEN 'active' THEN 0 WHEN 'paused' THEN 1 ELSE 2 END, l.created_at DESC
    ''').fetchall()
    employees = conn.execute(
        "SELECT id, name, arabic_name, employee_number FROM employees WHERE is_active = 1 "
        "ORDER BY CAST(employee_number AS INTEGER), employee_number").fetchall()
    return render_template('payroll_loans.html', rows=rows, employees=employees, now=datetime.now())


@payroll_bp.route('/payroll/loans/add', methods=['POST'])
@login_required
@require_permission('salary.calculate')
def add_loan():
    try:
        employee_id = int(request.form['employee_id'])
        principal = float(request.form['principal'])
        installment = float(request.form['monthly_installment'])
        start_month = int(request.form['start_month'])
        start_year = int(request.form['start_year'])
        reason = request.form.get('reason', '').strip()
        if principal <= 0 or installment <= 0 or installment > principal:
            flash(gettext('x.f_loan_amounts_invalid'), 'error')
            return redirect(url_for('payroll.loans'))
        conn = get_db_connection()
        conn.execute('''
            INSERT INTO employee_loans (employee_id, principal, monthly_installment, start_month, start_year, reason, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (employee_id, principal, installment, start_month, start_year, reason, _uid()))
        conn.commit()
        flash(gettext('x.f_loan_added'), 'success')
    except Exception as e:
        flash(gettext('x.f_loan_error') % {'p0': f'{str(e)}'}, 'error')
    return redirect(url_for('payroll.loans'))


@payroll_bp.route('/payroll/loans/pay/<int:loan_id>', methods=['POST'])
@login_required
@require_permission('salary.calculate')
def add_loan_payment(loan_id):
    conn = get_db_connection()
    loan = conn.execute('''
        SELECT l.*, COALESCE((SELECT SUM(amount) FROM loan_payments p WHERE p.loan_id = l.id), 0) AS paid
        FROM employee_loans l WHERE l.id = ?''', (loan_id,)).fetchone()
    if not loan or loan['status'] == 'settled':
        flash(gettext('x.f_not_found'), 'error')
        return redirect(url_for('payroll.loans'))
    try:
        amount = float(request.form['amount'])
        month = int(request.form['month'])
        year = int(request.form['year'])
        remaining = loan['principal'] - loan['paid']
        if amount <= 0 or amount > remaining + 0.0001:
            flash(gettext('x.f_payment_exceeds') % {'p0': f'{remaining:.3f}'}, 'error')
            return redirect(url_for('payroll.loans'))
        conn.execute('''INSERT INTO loan_payments (loan_id, month, year, amount, source, created_by)
                        VALUES (?, ?, ?, ?, 'manual', ?)''', (loan_id, month, year, amount, _uid()))
        if abs((loan['paid'] + amount) - loan['principal']) < 0.0001:
            conn.execute("UPDATE employee_loans SET status = 'settled' WHERE id = ?", (loan_id,))
        conn.commit()
        flash(gettext('x.f_payment_added'), 'success')
    except Exception as e:
        flash(gettext('x.f_loan_error') % {'p0': f'{str(e)}'}, 'error')
    return redirect(url_for('payroll.loans'))


@payroll_bp.route('/payroll/loans/status/<int:loan_id>', methods=['POST'])
@login_required
@require_permission('salary.calculate')
def set_loan_status(loan_id):
    new_status = request.form.get('status')
    if new_status not in ('active', 'paused', 'settled'):
        flash(gettext('x.f_not_found'), 'error')
        return redirect(url_for('payroll.loans'))
    conn = get_db_connection()
    conn.execute('UPDATE employee_loans SET status = ? WHERE id = ?', (new_status, loan_id))
    conn.commit()
    flash(gettext('x.f_loan_status_updated'), 'success')
    return redirect(url_for('payroll.loans'))


# ================= Employee Salary Structure APIs =================
@payroll_bp.route('/api/employees/salary_items')
@login_required
@require_permission('employee.view')
def get_salary_items():
    employee_id = request.args.get('employee_id')
    conn = get_db_connection()
    rows = conn.execute('''
        SELECT s.*, it.code, it.name_ar, it.name_en, it.kind
        FROM employee_salary_items s
        JOIN payroll_item_types it ON it.id = s.item_type_id
        WHERE s.employee_id = ? AND s.is_active = 1
        ORDER BY (s.end_date IS NOT NULL), it.kind, it.sort_order
    ''', (employee_id,)).fetchall()
    types = conn.execute('''
        SELECT id, code, name_ar, name_en, kind FROM payroll_item_types
        WHERE species = 'fixed' AND is_active = 1 ORDER BY kind, sort_order
    ''').fetchall()
    return jsonify({'success': True,
                    'items': [dict(r) for r in rows],
                    'types': [dict(t) for t in types]})


@payroll_bp.route('/api/employees/salary_items', methods=['POST'])
@login_required
@require_permission('employee.edit')
def add_salary_item():
    try:
        data = request.get_json()
        employee_id = int(data['employee_id'])
        item_type_id = int(data['item_type_id'])
        calc_method = data.get('calc_method', 'fixed')
        amount = float(data.get('amount') or 0)
        percent = float(data.get('percent_value') or 0)
        start_date = (data.get('start_date') or '').strip()
        end_date = (data.get('end_date') or '').strip() or None
        if calc_method not in ('fixed', 'percent_of_basic'):
            return jsonify({'success': False, 'message': 'bad method'}), 400
        if (calc_method == 'fixed' and amount <= 0) or (calc_method == 'percent_of_basic' and percent <= 0):
            return jsonify({'success': False, 'message': gettext('x.f_amount_positive')}), 400
        try:
            datetime.strptime(start_date, '%Y-%m-%d')
            if end_date:
                datetime.strptime(end_date, '%Y-%m-%d')
                if end_date < start_date:
                    return jsonify({'success': False,
                                    'message': gettext('x.f_end_before_start')}), 400
        except ValueError:
            return jsonify({'success': False,
                            'message': gettext('x.f_start_date_required')}), 400
        conn = get_db_connection()
        it = conn.execute("SELECT id FROM payroll_item_types WHERE id = ? AND species = 'fixed' AND is_active = 1",
                          (item_type_id,)).fetchone()
        if not it:
            return jsonify({'success': False, 'message': gettext('x.f_item_type_invalid')}), 400
        dup = conn.execute('''SELECT id FROM employee_salary_items
                              WHERE employee_id = ? AND item_type_id = ? AND is_active = 1''',
                           (employee_id, item_type_id)).fetchone()
        if dup:
            return jsonify({'success': False, 'message': gettext('x.f_item_duplicate')}), 400
        conn.execute('''INSERT INTO employee_salary_items
                        (employee_id, item_type_id, calc_method, amount, percent_value,
                         start_date, end_date, created_by)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                     (employee_id, item_type_id, calc_method, amount, percent,
                      start_date, end_date, _uid()))
        try:
            from utils.payroll_engine import log_employee_event
            _it = conn.execute('SELECT name_ar, name_en FROM payroll_item_types '
                               'WHERE id = ?', (item_type_id,)).fetchone()
            _label = (_it['name_ar'] or _it['name_en']) if _it else str(item_type_id)
            _val = (f'{percent}%' if calc_method == 'percent_of_basic' else str(amount))
            log_employee_event(conn, employee_id, _label, '', _val, 'pay',
                               user_id=_uid(), source='salary_item_add',
                               note=gettext('x.aud_from') + ' ' + str(start_date))
        except Exception as _e:
            print(f'audit log failed on salary item add: {_e}')
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@payroll_bp.route('/api/employees/salary_items/<int:id>', methods=['DELETE'])
@login_required
@require_permission('employee.edit')
def delete_salary_item(id):
    """Hard removal for a genuine entry mistake only. Prefer /end for the
    normal case of stopping a benefit going forward — hard-delete erases
    the item from every period, past included."""
    conn = get_db_connection()
    conn.execute('UPDATE employee_salary_items SET is_active = 0 WHERE id = ?', (id,))
    conn.commit()
    return jsonify({'success': True})


@payroll_bp.route('/api/employees/salary_items/<int:id>/end', methods=['POST'])
@login_required
@require_permission('employee.edit')
def end_salary_item(id):
    """Stop a fixed item from a given date forward WITHOUT touching any
    period up to and including that date."""
    conn = get_db_connection()
    row = conn.execute('SELECT * FROM employee_salary_items WHERE id = ? AND is_active = 1',
                       (id,)).fetchone()
    if not row:
        return jsonify({'success': False, 'code': 'not_found'}), 404
    data = request.get_json(silent=True) or {}
    end_date = (data.get('end_date') or '').strip()
    reason = (data.get('reason') or '').strip()
    try:
        datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'success': False, 'message': gettext('x.f_end_date_required')}), 400
    if row['start_date'] and end_date < row['start_date']:
        return jsonify({'success': False, 'message': gettext('x.f_end_before_start')}), 400
    if len(reason) < 3:
        return jsonify({'success': False, 'code': 'reason_required'}), 400
    conn.execute("UPDATE employee_salary_items SET end_date = ?, notes = ? WHERE id = ?",
                 (end_date, reason, id))
    try:
        from utils.payroll_engine import log_employee_event
        _it = conn.execute('''SELECT it.name_ar, it.name_en, si.employee_id,
                                     si.amount, si.percent_value, si.calc_method
                              FROM employee_salary_items si
                              JOIN payroll_item_types it ON it.id = si.item_type_id
                              WHERE si.id = ?''', (id,)).fetchone()
        if _it:
            _label = _it['name_ar'] or _it['name_en']
            _val = (f"{_it['percent_value']}%"
                    if _it['calc_method'] == 'percent_of_basic' else str(_it['amount']))
            log_employee_event(conn, _it['employee_id'], _label, _val, '', 'pay',
                               user_id=_uid(), source='salary_item_end',
                               note=f'{end_date} — {reason}')
    except Exception as _e:
        print(f'audit log failed on salary item end: {_e}')
    conn.commit()
    return jsonify({'success': True})


# ================= Monthly Payroll Sheet =================
@payroll_bp.route('/payroll/monthly')
@login_required
@require_permission('salary.calculate')
def monthly_sheet():
    now = datetime.now()
    month = int(request.args.get('month', now.month))
    year = int(request.args.get('year', now.year))
    return render_template('payroll_monthly.html', month=month, year=year)


@payroll_bp.route('/api/payroll/monthly')
@login_required
@require_permission('salary.calculate')
def api_monthly_payroll():
    from utils.payroll_engine import compute_monthly_payroll
    try:
        now = datetime.now()
        month = int(request.args.get('month', now.month))
        year = int(request.args.get('year', now.year))
        if not (1 <= month <= 12) or not (2000 <= year <= 2100):
            return jsonify({'success': False, 'message': 'bad period'}), 400
        conn = get_db_connection()
        data = compute_monthly_payroll(conn, month, year)
        saved = conn.execute('''
            SELECT r.id, r.status, r.created_at, r.employees_count, r.total_net,
                   r.period_start, r.period_end, r.approved_at,
                   u.username AS created_by_name,
                   ua.username AS approved_by_name
            FROM payroll_runs r
            LEFT JOIN users u ON u.id = r.created_by
            LEFT JOIN users ua ON ua.id = r.approved_by
            WHERE r.month = ? AND r.year = ?''', (month, year)).fetchone()
        return jsonify({'success': True, 'data': data,
                        'saved_run': dict(saved) if saved else None})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@payroll_bp.route('/api/payroll/monthly/save', methods=['POST'])
@login_required
@require_permission('salary.calculate')
def api_monthly_payroll_save():
    from utils.payroll_engine import save_monthly_payroll, HoursNotApprovedError, RunLockedError
    conn = get_db_connection()
    try:
        payload = request.get_json(silent=True) or {}
        month = int(payload.get('month'))
        year = int(payload.get('year'))
        if not (1 <= month <= 12) or not (2000 <= year <= 2100):
            return jsonify({'success': False, 'message': 'bad period'}), 400
        result = save_monthly_payroll(conn, month, year, _uid())
        conn.commit()
        return jsonify({'success': True, **result})
    except RunLockedError:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'code': 'run_locked'}), 409
    except HoursNotApprovedError as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'code': 'hours_not_approved',
                        'missing': e.missing}), 409
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'message': str(e)}), 500

# ================= Monthly Hours Approval (attestation model) =================
# Numbers are never hand-entered: this page ATTESTS server-computed metrics.
# Corrections flow only through documented instruments (excuse / sick / unpaid
# leave / manual punch), each recorded with creator and reason.

@payroll_bp.route('/payroll/hours_approval')
@login_required
@require_permission('salary.approve_hours')
def hours_approval():
    now = datetime.now()
    month = int(request.args.get('month', now.month))
    year = int(request.args.get('year', now.year))
    conn = get_db_connection()
    depts = [r['department'] for r in conn.execute(
        "SELECT DISTINCT department FROM employees WHERE department IS NOT NULL AND department != '' ORDER BY department").fetchall()]
    return render_template('payroll_hours_approval.html', month=month, year=year, departments=depts)


@payroll_bp.route('/api/payroll/hours_approval')
@login_required
@require_permission('salary.approve_hours')
def api_hours_approval():
    from utils.payroll_engine import compute_month_metrics, get_hours_approvals, approval_state, resolve_period
    try:
        now = datetime.now()
        month = int(request.args.get('month', now.month))
        year = int(request.args.get('year', now.year))
        dept = request.args.get('dept', '').strip()
        if not (1 <= month <= 12) or not (2000 <= year <= 2100):
            return jsonify({'success': False, 'message': 'bad period'}), 400

        conn = get_db_connection()
        from utils.payroll_engine import fetch_payroll_employees
        from utils.payroll_engine import resolve_period as _rp
        employees = fetch_payroll_employees(conn, dept=dept,
                                            period=_rp(conn, month, year))

        p_start, p_end = resolve_period(conn, month, year)
        metrics = compute_month_metrics(conn, month, year, employees)
        approvals = get_hours_approvals(conn, month, year)
        payroll_saved = conn.execute(
            'SELECT id FROM payroll_runs WHERE month = ? AND year = ?',
            (month, year)).fetchone()

        rows = []
        counts = {'approved': 0, 'stale': 0, 'pending': 0}
        for e in employees:
            m = metrics.get(e['id'], {})
            a = approvals.get(e['id'])
            state = approval_state(a, m, p_start.isoformat(), p_end.isoformat())
            counts[state] += 1
            rows.append({
                'employee_id': e['id'], 'employee_number': e['employee_number'],
                'name': e['name'], 'arabic_name': e['arabic_name'],
                'department': e['department'] or '',
                'metrics': m,
                'state': state,
                'approval': ({'notes': a['notes'],
                              'approved_by_name': a['approved_by_name'],
                              'approved_at': a['approved_at']} if a else None)
            })
        return jsonify({'success': True, 'rows': rows, 'counts': counts,
                        'total': len(rows),
                        'payroll_saved': bool(payroll_saved),
                        'period': {'start': p_start.isoformat(),
                                   'end': p_end.isoformat()}})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@payroll_bp.route('/api/payroll/hours_approval/approve', methods=['POST'])
@login_required
@require_permission('salary.approve_hours')
def api_hours_approve():
    from utils.payroll_engine import attest_hours
    conn = get_db_connection()
    try:
        payload = request.get_json(silent=True) or {}
        month = int(payload.get('month'))
        year = int(payload.get('year'))
        items = payload.get('rows') or []
        if not (1 <= month <= 12) or not (2000 <= year <= 2100) or not items:
            return jsonify({'success': False, 'message': 'bad payload'}), 400
        from utils.payroll_engine import is_period_locked
        if is_period_locked(conn, month, year):
            return jsonify({'success': False, 'code': 'period_closed',
                            'message': gettext('x.err_period_closed')}), 409
        # Attestation only: any numeric fields sent by the client are IGNORED.
        clean = [{'employee_id': int(i['employee_id']),
                  'notes': (i.get('notes') or '')} for i in items]
        saved = attest_hours(conn, month, year, clean, _uid())
        conn.commit()
        return jsonify({'success': True, 'saved': saved})
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'message': str(e)}), 500


@payroll_bp.route('/api/payroll/hours_approval/unapprove', methods=['POST'])
@login_required
@require_permission('salary.approve_hours')
def api_hours_unapprove():
    conn = get_db_connection()
    try:
        payload = request.get_json(silent=True) or {}
        month = int(payload.get('month'))
        year = int(payload.get('year'))
        emp_id = int(payload.get('employee_id'))
        from utils.payroll_engine import is_period_locked
        if is_period_locked(conn, month, year):
            return jsonify({'success': False, 'code': 'period_closed',
                            'message': gettext('x.err_period_closed')}), 409
        conn.execute('''DELETE FROM payroll_hours_approvals
                        WHERE employee_id = ? AND month = ? AND year = ?''',
                     (emp_id, month, year))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'message': str(e)}), 500


@payroll_bp.route('/api/payroll/hours_approval/employee_days')
@login_required
@require_permission('salary.approve_hours')
def api_hours_employee_days():
    """Day-by-day breakdown for ONE employee — the audit drill-down."""
    from utils.payroll_engine import compute_employee_days
    try:
        data = compute_employee_days(get_db_connection(),
                                     int(request.args.get('employee_id')),
                                     int(request.args.get('month')),
                                     int(request.args.get('year')))
        if not data:
            return jsonify({'success': False, 'message': 'not found'}), 404
        return jsonify({'success': True, **data})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ---------- Documented correction instruments ----------

@payroll_bp.route('/api/payroll/day_actions/excuse', methods=['POST'])
@login_required
@require_permission('salary.approve_hours')
def day_action_excuse():
    conn = get_db_connection()
    try:
        p = request.get_json(silent=True) or {}
        emp_id = int(p['employee_id'])
        d = str(p['date'])[:10]
        reason = (p.get('reason') or '').strip()
        side = p.get('side') or None
        datetime.strptime(d, '%Y-%m-%d')
        from utils.payroll_engine import date_period_locked
        if date_period_locked(conn, d):
            return jsonify({'success': False, 'code': 'period_closed',
                            'message': gettext('x.err_period_closed')}), 409
        if len(reason) < 3:
            return jsonify({'success': False, 'code': 'reason_required'}), 400
        if side not in (None, 'in', 'out'):
            return jsonify({'success': False, 'message': 'bad side'}), 400
        punches = conn.execute('''SELECT COUNT(*) AS n FROM attendance_records
                                  WHERE employee_id = ? AND DATE(check_time) = ?''',
                               (emp_id, d)).fetchone()['n']
        if punches > 0 and side is None:
            # A day with punches was worked: the excuse must target a punch side
            return jsonify({'success': False, 'code': 'side_required'}), 400
        if punches == 0:
            side = None  # whole-day excuse; a side is meaningless without punches
        dup = conn.execute('SELECT id FROM attendance_excuses WHERE employee_id = ? AND date = ?',
                           (emp_id, d)).fetchone()
        if dup:
            return jsonify({'success': False, 'code': 'already_excused'}), 400
        conn.execute('''INSERT INTO attendance_excuses (employee_id, date, reason, type, side, created_by)
                        VALUES (?, ?, ?, 'mission', ?, ?)''', (emp_id, d, reason, side, _uid()))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'message': str(e)}), 500


@payroll_bp.route('/api/payroll/day_actions/leave', methods=['POST'])
@login_required
@require_permission('salary.approve_hours')
def day_action_leave():
    conn = get_db_connection()
    try:
        p = request.get_json(silent=True) or {}
        emp_id = int(p['employee_id'])
        d = str(p['date'])[:10]
        kind = p.get('kind')
        reason = (p.get('reason') or '').strip()
        datetime.strptime(d, '%Y-%m-%d')
        from utils.payroll_engine import date_period_locked
        if date_period_locked(conn, d):
            return jsonify({'success': False, 'code': 'period_closed',
                            'message': gettext('x.err_period_closed')}), 409
        if kind not in ('sick', 'unpaid'):
            return jsonify({'success': False, 'message': 'bad kind'}), 400
        if len(reason) < 3:
            return jsonify({'success': False, 'code': 'reason_required'}), 400
        punches = conn.execute('''SELECT COUNT(*) AS n FROM attendance_records
                                  WHERE employee_id = ? AND DATE(check_time) = ?''',
                               (emp_id, d)).fetchone()['n']
        if punches > 0:
            # He was AT WORK that day: a full-day leave would contradict the
            # punches. Use a punch-side excuse instead.
            return jsonify({'success': False, 'code': 'has_punches'}), 400
        overlap = conn.execute('''SELECT id FROM leave_requests
                                  WHERE employee_id = ? AND status = 'approved'
                                    AND DATE(start_date) <= ? AND DATE(end_date) >= ?''',
                               (emp_id, d, d)).fetchone()
        if overlap:
            return jsonify({'success': False, 'code': 'leave_exists'}), 400
        if kind == 'sick':
            lt = conn.execute("SELECT id FROM leave_types WHERE name = 'إجازة مرضية'").fetchone() \
                 or conn.execute("SELECT id FROM leave_types WHERE name LIKE '%مرض%' OR LOWER(name) LIKE '%sick%'").fetchone()
            is_paid = 1
        else:
            lt = conn.execute("SELECT id FROM leave_types WHERE name = 'إجازة بدون راتب'").fetchone() \
                 or conn.execute("SELECT id FROM leave_types WHERE COALESCE(is_paid,1) = 0").fetchone()
            is_paid = 0
        if not lt:
            return jsonify({'success': False, 'code': 'leave_type_missing'}), 400
        conn.execute('''INSERT INTO leave_requests
                        (employee_id, leave_type_id, start_date, end_date, days_count,
                         leave_duration_type, reason, status, current_step,
                         is_paid_leave, is_deleted, updated_at)
                        VALUES (?, ?, ?, ?, 1, 'full_day', ?, 'approved', 1, ?, 0,
                                CURRENT_TIMESTAMP)''',
                     (emp_id, lt['id'], d, d, reason, is_paid))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'message': str(e)}), 500


# ================= Final Run Approval (lock) =================
@payroll_bp.route('/api/payroll/monthly/lock', methods=['POST'])
@login_required
@require_permission('salary.approve_run')
def api_run_lock():
    from utils.payroll_engine import lock_run, RunLockedError, post_to_ledger
    conn = get_db_connection()
    try:
        p = request.get_json(silent=True) or {}
        month = int(p.get('month'))
        year = int(p.get('year'))
        run_id = lock_run(conn, month, year, _uid())
        note = None
        prev_n = conn.execute('SELECT COUNT(*) AS n FROM payroll_ledger_postings '
                              'WHERE month = ? AND year = ?', (month, year)).fetchone()['n']
        if prev_n:
            r = conn.execute('SELECT unlock_reason FROM payroll_runs WHERE id = ?',
                             (run_id,)).fetchone()
            note = gettext('x.lg_note_relock') + ((r['unlock_reason'] or '') if r else '')
        post_to_ledger(conn, month, year, _uid(), note)
        conn.commit()
        return jsonify({'success': True})
    except RunLockedError:
        conn.rollback()
        return jsonify({'success': False, 'code': 'run_locked'}), 409
    except ValueError as e:
        conn.rollback()
        return jsonify({'success': False, 'code': str(e)}), 400
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'message': str(e)}), 500


@payroll_bp.route('/api/payroll/monthly/unlock', methods=['POST'])
@login_required
@require_permission('salary.approve_run')
def api_run_unlock():
    from utils.payroll_engine import unlock_run
    conn = get_db_connection()
    try:
        p = request.get_json(silent=True) or {}
        unlock_run(conn, int(p.get('month')), int(p.get('year')), _uid(),
                   p.get('reason') or '')
        conn.commit()
        return jsonify({'success': True})
    except ValueError as e:
        conn.rollback()
        return jsonify({'success': False, 'code': str(e)}), 400
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'message': str(e)}), 500


# ================= Print: all-employees attendance report =================
@payroll_bp.route('/payroll/hours_approval/print')
@login_required
@require_permission('salary.approve_hours')
def hours_approval_print():
    from utils.payroll_engine import (compute_month_metrics, get_hours_approvals,
                                      approval_state, resolve_period,
                                      compute_employee_days, fetch_payroll_employees)
    from flask_babel import gettext as _
    now = datetime.now()
    month = int(request.args.get('month', now.month))
    year = int(request.args.get('year', now.year))
    dept = request.args.get('dept', '').strip()

    conn = get_db_connection()
    from utils.payroll_engine import resolve_period as _rp
    employees = fetch_payroll_employees(conn, dept=dept,
                                        period=_rp(conn, month, year))

    p_start, p_end = resolve_period(conn, month, year)
    metrics = compute_month_metrics(conn, month, year, employees)
    approvals = get_hours_approvals(conn, month, year)

    blocks = []
    for e in employees:
        dd = compute_employee_days(conn, e['id'], month, year)
        if not dd:
            continue
        a = approvals.get(e['id'])
        m = metrics.get(e['id'], {})
        blocks.append({'employee': dd['employee'], 'days': dd['days'],
                       'presence_required': dd.get('presence_required', False),
                       'metrics': m,
                       'state': approval_state(a, m, p_start.isoformat(), p_end.isoformat()),
                       'approval': ({'by': a['approved_by_name'],
                                     'at': (a['approved_at'] or '')[:16],
                                     'notes': a['notes']} if a else None)})

    status_labels = {
        'present': _('x.ds_present'), 'partial': _('x.ds_partial'),
        'absent': _('x.ds_absent'), 'weekly_off': _('x.ds_weekly_off'),
        'holiday': _('x.ds_holiday'), 'leave_sick': _('x.ds_sick'),
        'leave_unpaid': _('x.ds_unpaid'), 'leave_paid': _('x.ds_leave'),
        'excused': _('x.ds_excused'), 'not_hired': _('x.ds_not_hired'),
    }
    day_names = [_('day.sun'), _('day.mon'), _('day.tue'), _('day.wed'),
                 _('day.thu'), _('day.fri'), _('day.sat')]
    return render_template('payroll_hours_print.html',
                           blocks=blocks, month=month, year=year,
                           period={'start': p_start.isoformat(), 'end': p_end.isoformat()},
                           status_labels=status_labels, day_names=day_names,
                           printed_at=now.strftime('%Y-%m-%d %H:%M'))


@payroll_bp.route('/api/payroll/day_actions/hourly_perm', methods=['POST'])
@login_required
@require_permission('salary.approve_hours')
def day_action_hourly_perm():
    """Hourly permission (إذن ساعات): recorded into the EXISTING leave system
    as an approved hourly leave row. Limits come from settings."""
    from utils.payroll_engine import resolve_period, _t2m
    from utils.settings_utils import get_salary_settings_v2
    conn = get_db_connection()
    try:
        p = request.get_json(silent=True) or {}
        emp_id = int(p['employee_id'])
        d = str(p['date'])[:10]
        f_t = str(p.get('from') or '')[:5]
        t_t = str(p.get('to') or '')[:5]
        reason = (p.get('reason') or '').strip()
        datetime.strptime(d, '%Y-%m-%d')
        from utils.payroll_engine import date_period_locked
        if date_period_locked(conn, d):
            return jsonify({'success': False, 'code': 'period_closed',
                            'message': gettext('x.err_period_closed')}), 409
        if len(reason) < 3:
            return jsonify({'success': False, 'code': 'reason_required'}), 400
        hours = (_t2m(t_t, '00:00') - _t2m(f_t, '00:00')) / 60.0
        if not f_t or not t_t or hours <= 0:
            return jsonify({'success': False, 'code': 'hp_bad_times'}), 400

        sal = get_salary_settings_v2(conn)
        max_h = float(sal.get('hourly_perm_max_hours_per_day', 2) or 2)
        max_d = int(float(sal.get('hourly_perm_max_days_per_month', 4) or 4))
        if hours > max_h + 1e-9:
            return jsonify({'success': False, 'code': 'hp_hours',
                            'limit': max_h}), 400

        punches = conn.execute('''SELECT COUNT(*) AS n FROM attendance_records
                                  WHERE employee_id = ? AND DATE(check_time) = ?''',
                               (emp_id, d)).fetchone()['n']
        if punches == 0:
            return jsonify({'success': False, 'code': 'hp_nopunch'}), 400

        dup = conn.execute('''SELECT id FROM leave_requests
                              WHERE employee_id = ? AND status = 'approved'
                                AND COALESCE(is_deleted,0) = 0
                                AND leave_duration_type = 'hourly'
                                AND DATE(start_date) = ?''', (emp_id, d)).fetchone()
        if dup:
            return jsonify({'success': False, 'code': 'hp_dup'}), 400

        overlap = conn.execute('''SELECT id FROM leave_requests
                                  WHERE employee_id = ? AND status = 'approved'
                                    AND COALESCE(is_deleted,0) = 0
                                    AND COALESCE(leave_duration_type,'full_day') != 'hourly'
                                    AND NOT (DATE(end_date) < ? OR DATE(start_date) > ?)''',
                               (emp_id, d, d)).fetchone()
        if overlap:
            return jsonify({'success': False, 'code': 'leave_exists'}), 400

        dt = datetime.strptime(d, '%Y-%m-%d').date()
        # month = the payroll period containing this date (labeled month resolution)
        probe = dt.replace(day=1)
        for cand in ((probe.month, probe.year),
                     (1 if probe.month == 12 else probe.month + 1,
                      probe.year + 1 if probe.month == 12 else probe.year)):
            ps, pe = resolve_period(conn, cand[0], cand[1])
            if ps <= dt <= pe:
                break
        used = conn.execute('''SELECT COUNT(DISTINCT DATE(start_date)) AS n
                               FROM leave_requests
                               WHERE employee_id = ? AND status = 'approved'
                                 AND COALESCE(is_deleted,0) = 0
                                 AND leave_duration_type = 'hourly'
                                 AND DATE(start_date) BETWEEN ? AND ?''',
                            (emp_id, ps.isoformat(), pe.isoformat())).fetchone()['n']
        if used >= max_d:
            return jsonify({'success': False, 'code': 'hp_limit', 'limit': max_d}), 400

        lt = conn.execute('''SELECT id FROM leave_types
                             WHERE is_hourly_permission = 1
                               AND COALESCE(is_deleted,0) = 0
                             ORDER BY id LIMIT 1''').fetchone()
        if not lt:
            return jsonify({'success': False, 'code': 'leave_type_missing'}), 400

        conn.execute('''INSERT INTO leave_requests
                        (employee_id, leave_type_id, start_date, end_date, days_count,
                         leave_duration_type, start_time, end_time, reason, status,
                         current_step, is_paid_leave, is_deleted, created_by, updated_at)
                        VALUES (?, ?, ?, ?, 0, 'hourly', ?, ?, ?, 'approved', 1, 1, 0, ?,
                                CURRENT_TIMESTAMP)''',
                     (emp_id, lt['id'], d, d, f_t, t_t, reason, session.get('user_id')))
        conn.commit()
        return jsonify({'success': True, 'hours': round(hours, 2)})
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'message': str(e)}), 500


@payroll_bp.route('/payroll/hourly_perms')
@login_required
@require_permission('salary.approve_hours')
def hourly_perms_page():
    from utils.payroll_engine import resolve_period
    from utils.settings_utils import get_salary_settings_v2
    conn = get_db_connection()
    today = datetime.now().date()
    for cand in ((today.month, today.year),
                 (1 if today.month == 12 else today.month + 1,
                  today.year + 1 if today.month == 12 else today.year)):
        ps, pe = resolve_period(conn, cand[0], cand[1])
        if ps <= today <= pe:
            break
    from utils.payroll_engine import fetch_payroll_employees as _fpe
    emps = _fpe(conn, period=(ps, pe))
    sal = get_salary_settings_v2(conn)
    return render_template('payroll_hourly_perms.html',
                           employees=emps, month=cand[0], year=cand[1],
                           max_h=sal.get('hourly_perm_max_hours_per_day', '2'),
                           max_d=sal.get('hourly_perm_max_days_per_month', '4'))


@payroll_bp.route('/api/payroll/hourly_perms')
@login_required
@require_permission('salary.approve_hours')
def api_hourly_perms_list():
    from utils.payroll_engine import resolve_period, _t2m
    conn = get_db_connection()
    try:
        month = int(request.args.get('month'))
        year = int(request.args.get('year'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'bad period'}), 400
    dept = request.args.get('dept', '').strip()
    p_start, p_end = resolve_period(conn, month, year)
    q = """SELECT lr.id, lr.employee_id, lr.start_date, lr.start_time, lr.end_time,
                  lr.reason, e.employee_number, e.name, e.arabic_name, e.department,
                  u.username AS by_name
           FROM leave_requests lr
           JOIN employees e ON e.id = lr.employee_id
           LEFT JOIN users u ON u.id = lr.created_by
           WHERE lr.leave_duration_type = 'hourly'
             AND lr.status = 'approved'
             AND COALESCE(lr.is_deleted, 0) = 0
             AND DATE(lr.start_date) BETWEEN ? AND ?"""
    params = [p_start.isoformat(), p_end.isoformat()]
    if dept:
        q += ' AND e.department = ?'
        params.append(dept)
    q += ' ORDER BY DATE(lr.start_date), CAST(e.employee_number AS INTEGER)'
    rows = []
    tot_h = 0.0
    for r in conn.execute(q, params).fetchall():
        h = max(0.0, (_t2m(r['end_time'], '00:00') - _t2m(r['start_time'], '00:00')) / 60.0)
        tot_h += h
        rows.append({'id': r['id'], 'employee_id': r['employee_id'],
                     'employee_number': r['employee_number'],
                     'name': r['name'], 'arabic_name': r['arabic_name'],
                     'department': r['department'],
                     'date': str(r['start_date'])[:10],
                     'from': str(r['start_time'])[:5], 'to': str(r['end_time'])[:5],
                     'hours': round(h, 2), 'reason': r['reason'],
                     'by': r['by_name']})
    return jsonify({'success': True, 'rows': rows,
                    'period': {'start': p_start.isoformat(), 'end': p_end.isoformat()},
                    'totals': {'count': len(rows), 'hours': round(tot_h, 2)}})


@payroll_bp.route('/api/payroll/hourly_perms/delete', methods=['POST'])
@login_required
@require_permission('salary.approve_hours')
def api_hourly_perm_delete():
    conn = get_db_connection()
    try:
        pid = int((request.get_json(silent=True) or {}).get('id'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'bad id'}), 400
    row = conn.execute("""SELECT id, start_date FROM leave_requests
                          WHERE id = ? AND leave_duration_type = 'hourly'
                            AND COALESCE(is_deleted, 0) = 0""", (pid,)).fetchone()
    if not row:
        return jsonify({'success': False, 'code': 'not_found'}), 404
    from utils.payroll_engine import date_period_locked
    if date_period_locked(conn, str(row['start_date'])[:10]):
        return jsonify({'success': False, 'code': 'period_closed',
                        'message': gettext('x.err_period_closed')}), 409
    conn.execute("""UPDATE leave_requests SET is_deleted = 1,
                    updated_at = CURRENT_TIMESTAMP WHERE id = ?""", (pid,))
    conn.commit()
    return jsonify({'success': True})


@payroll_bp.route('/payroll/ledger')
@login_required
@require_permission('salary.view')
def ledger_page():
    return render_template('payroll_ledger.html')


@payroll_bp.route('/api/payroll/ledger')
@login_required
@require_permission('salary.view')
def api_ledger_list():
    conn = get_db_connection()
    out = []
    for p in conn.execute('SELECT p.*, u.username AS by_name '
                          'FROM payroll_ledger_postings p '
                          'LEFT JOIN users u ON u.id = p.posted_by '
                          'ORDER BY p.id DESC').fetchall():
        lines = [{'employee_number': l['employee_number'],
                  'name': l['name'], 'arabic_name': l['arabic_name'],
                  'basic': l['basic'], 'allow': l['total_allowances'],
                  'ded': l['total_deductions'], 'net': l['net']}
                 for l in conn.execute('SELECT * FROM payroll_ledger_lines '
                                       'WHERE posting_id = ? ORDER BY id',
                                       (p['id'],)).fetchall()]
        out.append({'id': p['id'], 'month': p['month'], 'year': p['year'],
                    'period_start': p['period_start'], 'period_end': p['period_end'],
                    'note': p['note'], 'employees_count': p['employees_count'],
                    'total_basic': p['total_basic'],
                    'total_allowances': p['total_allowances'],
                    'total_deductions': p['total_deductions'],
                    'total_net': p['total_net'],
                    'chain_hash': p['chain_hash'], 'posted_at': p['posted_at'],
                    'by': p['by_name'], 'lines': lines})
    return jsonify({'success': True, 'postings': out})


@payroll_bp.route('/api/payroll/ledger/verify')
@login_required
@require_permission('salary.view')
def api_ledger_verify():
    from utils.payroll_engine import verify_ledger
    ok, broken = verify_ledger(get_db_connection())
    return jsonify({'success': True, 'ok': ok, 'broken_posting_id': broken})


def _employee_history_data(conn, employee_id):
    import json as _j
    from utils.payroll_engine import fetch_payroll_employees
    emp_rows = fetch_payroll_employees(conn, employee_id=employee_id,
                                       active_only=False)
    if not emp_rows:
        return None
    emp = dict(emp_rows[0])
    months = {}
    for l in conn.execute(
            """SELECT p.month, p.year, p.period_start, p.period_end,
                      p.id AS posting_id, p.posted_at,
                      l.basic, l.total_allowances, l.total_deductions,
                      l.net, l.details_json
               FROM payroll_ledger_lines l
               JOIN payroll_ledger_postings p ON p.id = l.posting_id
               WHERE l.employee_id = ?
               ORDER BY p.id""", (employee_id,)).fetchall():
        months[(l['year'], l['month'])] = {
            'month': l['month'], 'year': l['year'],
            'period_start': l['period_start'], 'period_end': l['period_end'],
            'source': 'ledger', 'posting_id': l['posting_id'],
            'posted_at': l['posted_at'],
            'basic': l['basic'], 'allow': l['total_allowances'],
            'ded': l['total_deductions'], 'net': l['net'],
            'details': _j.loads(l['details_json'] or '{}')}
    for r in conn.execute(
            """SELECT r.month, r.year, r.period_start, r.period_end, r.status,
                      rl.basic, rl.total_allowances, rl.total_deductions,
                      rl.net, rl.details_json
               FROM payroll_run_lines rl
               JOIN payroll_runs r ON r.id = rl.run_id
               WHERE rl.employee_id = ?""", (employee_id,)).fetchall():
        key = (r['year'], r['month'])
        if key in months:
            continue   # the ledger posting is authoritative — never replaced
        months[key] = {
            'month': r['month'], 'year': r['year'],
            'period_start': r['period_start'], 'period_end': r['period_end'],
            'source': 'run', 'posting_id': None, 'posted_at': None,
            'basic': r['basic'],
            'allow': r['total_allowances'], 'ded': r['total_deductions'],
            'net': r['net'],
            'details': _j.loads(r['details_json'] or '{}')}
    out_months = [months[k] for k in sorted(months.keys(), reverse=True)]
    tot = {'net': round(sum(m['net'] or 0 for m in out_months
                            if m['source'] == 'ledger'), 3),
           'allow': round(sum(m['allow'] or 0 for m in out_months
                              if m['source'] == 'ledger'), 3),
           'ded': round(sum(m['ded'] or 0 for m in out_months
                            if m['source'] == 'ledger'), 3),
           'months': sum(1 for m in out_months if m['source'] == 'ledger')}
    from utils.payroll_engine import fetch_employee_loans_detail
    loans = fetch_employee_loans_detail(conn, employee_id)
    return {'employee': {'id': emp['id'],
                         'employee_number': emp['employee_number'],
                         'name': emp['name'], 'arabic_name': emp['arabic_name'],
                         'department': emp['department'],
                         'position': emp['position'],
                         'hire_date': str(emp['hire_date'] or '')[:10],
                         'end_of_service_date':
                             str(emp['end_of_service_date'] or '')[:10],
                         'is_active': emp['is_active'],
                         'basic': emp['salary']},
            'months': out_months, 'totals': tot, 'loans': loans}


@payroll_bp.route('/payroll/employee_history')
@login_required
@require_permission('salary.view')
def employee_history_page():
    from utils.payroll_engine import fetch_payroll_employees
    conn = get_db_connection()
    emps = fetch_payroll_employees(conn, active_only=False)
    sel = request.args.get('employee_id', type=int)
    return render_template('payroll_employee_history.html',
                           employees=emps, selected=sel)


@payroll_bp.route('/api/payroll/employee_history')
@login_required
@require_permission('salary.view')
def api_employee_history():
    conn = get_db_connection()
    emp_id = request.args.get('employee_id', type=int)
    if not emp_id:
        return jsonify({'success': False, 'message': 'employee_id required'}), 400
    data = _employee_history_data(conn, emp_id)
    if data is None:
        return jsonify({'success': False, 'code': 'not_found'}), 404
    return jsonify({'success': True, 'data': data})


@payroll_bp.route('/payroll/employee_history/print')
@login_required
@require_permission('salary.view')
def employee_history_print():
    conn = get_db_connection()
    emp_id = request.args.get('employee_id', type=int)
    data = _employee_history_data(conn, emp_id) if emp_id else None
    if data is None:
        return 'employee not found', 404
    from utils.settings_utils import get_system_settings, get_salary_settings_v2
    sysx = get_system_settings() or {}
    try:
        _dec = int(float(get_salary_settings_v2(conn)
                         .get('rounding_display_decimals', 2) or 2))
    except (TypeError, ValueError):
        _dec = 2
    _dec = min(max(_dec, 0), 3)
    return render_template('payroll_employee_history_print.html',
                           d=data, company=sysx.get('company_name', ''),
                           currency=sysx.get('currency_symbol', ''),
                           fmt='%.' + str(_dec) + 'f')


@payroll_bp.route('/api/payroll/item_types', methods=['POST'])
@login_required
@require_permission('employee.edit')
def create_item_type():
    """Create a new fixed-earning allowance type on the fly, for HR to add
    a benefit that is not in the predefined catalog."""
    import re as _re
    conn = get_db_connection()
    data = request.get_json(silent=True) or {}
    name_ar = (data.get('name_ar') or '').strip()
    name_en = (data.get('name_en') or '').strip()
    if len(name_ar) < 2 or len(name_en) < 2:
        return jsonify({'success': False, 'message': gettext('x.f_name_required')}), 400
    base = _re.sub(r'[^a-z0-9]+', '_', name_en.lower()).strip('_') or 'allowance'
    code = base
    n = 1
    while conn.execute("SELECT 1 FROM payroll_item_types WHERE code = ?", (code,)).fetchone():
        n += 1
        code = f'{base}_{n}'
    max_sort = conn.execute(
        "SELECT COALESCE(MAX(sort_order), 100) FROM payroll_item_types "
        "WHERE species = 'fixed'").fetchone()[0]
    cur = conn.execute("""INSERT INTO payroll_item_types
                          (code, name_ar, name_en, kind, species, affects,
                           is_system, sort_order)
                          VALUES (?, ?, ?, 'earning', 'fixed', 'monthly', 0, ?)""",
                       (code, name_ar, name_en, int(max_sort) + 1))
    conn.commit()
    return jsonify({'success': True, 'id': cur.lastrowid, 'code': code})


def _monthly_export_data(conn, month, year, dept=''):
    """Shared payload for the printable sheet and the Excel export, so both
    always show exactly what the on-screen payroll shows."""
    from utils.payroll_engine import compute_monthly_payroll, is_period_locked
    data = compute_monthly_payroll(conn, month, year)
    rows = data['rows']
    if dept:
        rows = [r for r in rows if (r.get('department') or '') == dept]
        data = dict(data)
        data['rows'] = rows
        t = {'basic': 0.0, 'allowances': 0.0, 'deductions': 0.0, 'net': 0.0}
        for r in rows:
            t['basic'] += r['basic'] or 0
            t['allowances'] += r['total_allowances'] or 0
            t['deductions'] += r['total_deductions'] or 0
            t['net'] += r['net'] or 0
        data['totals'] = {**data['totals'], **{k: round(v, 3) for k, v in t.items()}}
    data['locked'] = is_period_locked(conn, month, year)
    return data


@payroll_bp.route('/payroll/monthly/print')
@login_required
@require_permission('salary.calculate')
def monthly_print():
    now = datetime.now()
    month = int(request.args.get('month', now.month))
    year = int(request.args.get('year', now.year))
    dept = request.args.get('dept', '').strip()
    conn = get_db_connection()
    data = _monthly_export_data(conn, month, year, dept)
    from utils.settings_utils import get_system_settings, get_salary_settings_v2
    sysx = get_system_settings() or {}
    try:
        dec = int(float(get_salary_settings_v2(conn)
                        .get('rounding_display_decimals', 2) or 2))
    except (TypeError, ValueError):
        dec = 2
    dec = min(max(dec, 0), 3)
    return render_template('payroll_monthly_print.html', d=data, dept=dept,
                           company=sysx.get('company_name', ''),
                           currency=sysx.get('currency_symbol', ''),
                           fmt='%.' + str(dec) + 'f')


@payroll_bp.route('/payroll/monthly/export')
@login_required
@require_permission('salary.calculate')
def monthly_export():
    """Excel export with one row per employee and every earning/deduction as
    its own column, so the sheet can be pivoted or handed to accounting."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    now = datetime.now()
    month = int(request.args.get('month', now.month))
    year = int(request.args.get('year', now.year))
    dept = request.args.get('dept', '').strip()
    conn = get_db_connection()
    data = _monthly_export_data(conn, month, year, dept)
    rows = data['rows']

    ar = (str(get_locale()) == 'ar')

    def item_label(it):
        return (it.get('name_ar') or it.get('name_en')) if ar else \
               (it.get('name_en') or it.get('name_ar'))

    earn_cols, ded_cols = [], []
    for r in rows:
        for a in r['allowances']:
            lbl = item_label(a)
            if lbl not in earn_cols:
                earn_cols.append(lbl)
        for x in r['deductions']:
            lbl = item_label(x)
            if lbl not in ded_cols:
                ded_cols.append(lbl)

    wb = Workbook()
    ws = wb.active
    ws.title = f'{month:02d}-{year}'
    ws.sheet_view.rightToLeft = ar

    head = [gettext('x.hp_employee_no'), gettext('x.hp_employee'),
            gettext('x.department'), gettext('x.pm_basic')]
    head += earn_cols
    head += [gettext('x.pm_allowances')]
    head += ded_cols
    head += [gettext('x.pm_deductions'), gettext('x.pm_net')]

    title = f"{gettext('x.payroll_monthly')} — {month:02d}/{year}"
    if dept:
        title += f' — {dept}'
    ws.append([title])
    ws.append([f"{data['period']['start']} ← {data['period']['end']}"
               + ('  •  ' + gettext('x.pm_locked_badge') if data.get('locked') else '')])
    ws.append([])
    ws.append(head)

    header_row = 4
    fill = PatternFill('solid', fgColor='1F5C4A')
    thin = Side(style='thin', color='CCCCCC')
    for c in range(1, len(head) + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1).font = Font(size=10, color='666666')

    for r in rows:
        emap = {item_label(a): a['amount'] for a in r['allowances']}
        dmap = {item_label(x): x['amount'] for x in r['deductions']}
        line = [r.get('employee_number'), r.get('arabic_name') or r.get('name'),
                r.get('department') or '', r['basic']]
        line += [emap.get(c, 0) for c in earn_cols]
        line += [r['total_allowances']]
        line += [dmap.get(c, 0) for c in ded_cols]
        line += [r['total_deductions'], r['net']]
        ws.append(line)

    t = data['totals']
    total_line = [gettext('x.eh_totals'), '', '', t.get('basic', 0)]
    total_line += ['' for _ in earn_cols]
    total_line += [t.get('allowances', 0)]
    total_line += ['' for _ in ded_cols]
    total_line += [t.get('deductions', 0), t.get('net', 0)]
    ws.append(total_line)
    last = ws.max_row
    for c in range(1, len(head) + 1):
        cell = ws.cell(row=last, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='EFEFE9')

    for c in range(1, len(head) + 1):
        letter = get_column_letter(c)
        width = max(len(str(ws.cell(row=header_row, column=c).value or '')) + 4, 12)
        ws.column_dimensions[letter].width = min(width, 30)
        if c >= 4:
            for rr in range(header_row + 1, last + 1):
                ws.cell(row=rr, column=c).number_format = '#,##0.000'
    ws.freeze_panes = ws.cell(row=header_row + 1, column=4)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f'payroll_{year}_{month:02d}' + (f'_{dept}' if dept else '') + '.xlsx'
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
