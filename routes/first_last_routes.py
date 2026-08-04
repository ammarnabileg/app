# -*- coding: utf-8 -*-
"""مسارات تقرير الأول والأخير."""
import io
from datetime import date, timedelta

from flask import (Blueprint, render_template, request, jsonify,
                   send_file, redirect, url_for, flash)

from utils.db import get_db_connection
from utils.auth import login_required
from utils.rbac import require_permission
from utils.first_last_report import build_first_last

fl_bp = Blueprint('first_last', __name__)


def _args():
    today = date.today()
    start = request.args.get('start') or (today - timedelta(days=6)).isoformat()
    end = request.args.get('end') or today.isoformat()
    ids = request.args.getlist('employee_ids') or request.args.getlist('employee_ids[]')
    emp_ids = []
    for v in ids:
        for part in str(v).split(','):
            part = part.strip()
            if part.isdigit():
                emp_ids.append(int(part))
    return (start, end, emp_ids or None,
            request.args.get('department') or None,
            request.args.get('include_absent', '1') != '0')


@fl_bp.route('/reports/first-last')
@login_required
@require_permission('attendance.view')
def index():
    conn = get_db_connection()
    employees = [dict(r) for r in conn.execute(
        """SELECT id, employee_number, name, arabic_name, department
           FROM employees WHERE COALESCE(is_active, 1) = 1
           ORDER BY CAST(employee_number AS INTEGER)""").fetchall()]
    depts = [r[0] for r in conn.execute(
        "SELECT DISTINCT department FROM employees WHERE department <> ''").fetchall()]
    today = date.today()
    return render_template('first_last_report.html',
                           employees=employees, departments=depts,
                           default_start=(today - timedelta(days=6)).isoformat(),
                           default_end=today.isoformat())


@fl_bp.route('/api/reports/first-last')
@login_required
@require_permission('attendance.view')
def api():
    start, end, ids, dept, inc = _args()
    rep = build_first_last(get_db_connection(), start, end, ids, dept, inc)
    if rep is None:
        return jsonify({'success': False, 'message': 'مدى تواريخ غير صحيح'}), 400
    return jsonify({'success': True, 'data': rep})


@fl_bp.route('/reports/first-last/print')
@login_required
@require_permission('attendance.view')
def print_view():
    start, end, ids, dept, inc = _args()
    rep = build_first_last(get_db_connection(), start, end, ids, dept, inc)
    if rep is None:
        flash('مدى تواريخ غير صحيح', 'error')
        return redirect(url_for('first_last.index'))
    conn = get_db_connection()
    header = ''
    try:
        row = conn.execute(
            "SELECT setting_value FROM app_settings WHERE setting_key = 'report_header'"
        ).fetchone()
        header = row[0] if row else ''
    except Exception:
        pass
    return render_template('first_last_print.html', rep=rep, header=header)


@fl_bp.route('/reports/first-last/export')
@login_required
@require_permission('attendance.view')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    start, end, ids, dept, inc = _args()
    rep = build_first_last(get_db_connection(), start, end, ids, dept, inc)
    if rep is None:
        flash('مدى تواريخ غير صحيح', 'error')
        return redirect(url_for('first_last.index'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'First-Last'
    ws.sheet_view.rightToLeft = True

    thin = Side(style='thin', color='999999')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head = PatternFill('solid', fgColor='1F5C4A')
    hfont = Font(bold=True, color='FFFFFF', size=10)
    # الألوان نفسها المستعملة في العرض: من يقرأ الورقة بعد الطباعة يجد
    # الدلالة ذاتها بلا مفتاح منفصل.
    fill_single = PatternFill('solid', fgColor='F5E6E2')
    fill_multi = PatternFill('solid', fgColor='FFF3CD')
    fill_absent = PatternFill('solid', fgColor='EFEFEF')

    ws.cell(row=1, column=1, value='تقرير الأول والأخير').font = Font(bold=True, size=13)
    ws.cell(row=2, column=1,
            value=f"من {rep['start']} إلى {rep['end']}").font = Font(size=10)
    r = 4

    cols = ['رقم الموظف', 'الاسم', 'القسم', 'التاريخ', 'اليوم',
            'أول تسجيل دخول', 'آخر تسجيل خروج', 'الوقت الإجمالي', 'سريال الجهاز']
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=r, column=i, value=c)
        cell.font = hfont
        cell.fill = head
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    header_row = r
    r += 1

    for blk in rep['employees']:
        e = blk['employee']
        for d in blk['days']:
            vals = [e['employee_number'], e['arabic_name'] or e['name'],
                    e['department'] or '', d['date'], d['day_ar'],
                    d['first_in'] or '—', d['last_out'] or '—',
                    d['total'], d['device']]
            for i, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=i, value=v)
                cell.border = border
                if d['flag'] == 'single':
                    cell.fill = fill_single
                elif d['flag'] == 'multi_device':
                    cell.fill = fill_multi
                elif d['flag'] == 'absent':
                    cell.fill = fill_absent
            r += 1
        s = blk['summary']
        ws.cell(row=r, column=1,
                value=(f"الإجمالي — أيام الحضور: {s['present_days']} | "
                       f"أيام ناقصة انصراف: {s['missing_out_days']}")).font = Font(bold=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.cell(row=r, column=8, value=s['total']).font = Font(bold=True)
        r += 2

    g = rep['grand']
    ws.cell(row=r, column=1,
            value=(f"الإجمالي العام — {g['employee_count']} موظف | "
                   f"أيام حضور: {g['present_days']} | "
                   f"ناقص انصراف: {g['missing_out_days']}")).font = Font(bold=True, size=11)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=8, value=g['total']).font = Font(bold=True, size=11)

    for i, w in enumerate([14, 26, 16, 12, 10, 14, 14, 13, 18], start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"first_last_{rep['start']}_{rep['end']}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
