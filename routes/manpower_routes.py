# -*- coding: utf-8 -*-
"""مسارات كشف قوى عاملة المقاولات.

الصلاحية `report.manpower` مستقلّة عن صلاحيات الرواتب عمدًا: الكشف يُتاح
لعدد محدود من الشركات ولمستخدمين بعينهم، وربطه بصلاحية عامة يكشف مطالبات
مقاولٍ لمن لا يعنيه.
"""
import io
import os
from datetime import datetime

from flask import (Blueprint, render_template, request, jsonify,
                   send_file, redirect, url_for, flash, session)
from flask_babel import gettext

from utils.db import get_db_connection
from utils.auth import login_required
from utils.rbac import require_permission
from utils.manpower_report import build_report, list_contracts, get_contract

manpower_bp = Blueprint('manpower', __name__)

ALLOWED_LOGO = {'.png', '.jpg', '.jpeg', '.svg', '.webp'}


@manpower_bp.route('/reports/manpower')
@login_required
@require_permission('report.manpower')
def index():
    conn = get_db_connection()
    contracts = list_contracts(conn, active_only=False)
    today = datetime.now()
    return render_template('manpower_report.html',
                           contracts=contracts,
                           cur_month=today.month, cur_year=today.year)


@manpower_bp.route('/api/manpower/report')
@login_required
@require_permission('report.manpower')
def report_api():
    try:
        cid = int(request.args.get('contract_id', 0))
        month = int(request.args.get('month'))
        year = int(request.args.get('year'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'معاملات غير صحيحة'}), 400

    rep = build_report(get_db_connection(), cid, month, year,
                       department=request.args.get('department') or None)
    if not rep:
        return jsonify({'success': False, 'message': 'العقد غير موجود'}), 404
    return jsonify({'success': True, 'data': rep})


@manpower_bp.route('/reports/manpower/print')
@login_required
@require_permission('report.manpower')
def print_view():
    try:
        cid = int(request.args.get('contract_id', 0))
        month = int(request.args.get('month'))
        year = int(request.args.get('year'))
    except (TypeError, ValueError):
        flash('معاملات غير صحيحة', 'error')
        return redirect(url_for('manpower.index'))

    rep = build_report(get_db_connection(), cid, month, year,
                       department=request.args.get('department') or None)
    if not rep:
        flash('العقد غير موجود', 'error')
        return redirect(url_for('manpower.index'))
    return render_template('manpower_report_print.html', rep=rep)


@manpower_bp.route('/reports/manpower/export')
@login_required
@require_permission('report.manpower')
def export_excel():
    """تصدير بنفس ترتيب أعمدة كشف الوزارة.

    الترتيب مقصود ولا يُغيَّر: الكشف يُقدَّم إلى جهة تراجعه بعينها، وأي
    إزاحة عمود تجعل المراجعة تفشل.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    try:
        cid = int(request.args.get('contract_id', 0))
        month = int(request.args.get('month'))
        year = int(request.args.get('year'))
    except (TypeError, ValueError):
        flash('معاملات غير صحيحة', 'error')
        return redirect(url_for('manpower.index'))

    rep = build_report(get_db_connection(), cid, month, year,
                       department=request.args.get('department') or None)
    if not rep:
        flash('العقد غير موجود', 'error')
        return redirect(url_for('manpower.index'))

    c = rep['contract']
    wb = Workbook()
    ws = wb.active
    ws.title = f"{rep['month_en'][:3]}{year}"

    thin = Side(style='thin', color='999999')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill('solid', fgColor='1F5C4A')
    head_font = Font(bold=True, color='FFFFFF', size=9)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    r = 1
    for line in (c.get('ministry_line1'), c.get('station_name'),
                 f"Contract No.: {c.get('contract_no') or ''}",
                 c.get('work_description'),
                 f"Payment Ref.: {c.get('payment_ref') or ''}",
                 f"MANPOWER SUPPLY FOR {rep['month_en']} {year}"):
        if line:
            ws.cell(row=r, column=1, value=line).font = Font(bold=True, size=10)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=23)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')
            r += 1
    r += 1

    groups = [('', 3), ('ATTENDANCE', 4), ('NORMAL O.T.', 4),
              ('FRIDAY O.T.', 4), ('HOLIDAY O.T.', 4), ('', 2)]
    col = 1
    for label, span in groups:
        if label:
            ws.merge_cells(start_row=r, start_column=col,
                           end_row=r, end_column=col + span - 1)
            cell = ws.cell(row=r, column=col, value=label)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = center
        col += span

    r += 1
    headers = ['Item', 'SI. No.', 'Name', 'Designation',
               'Per month Salary', 'Per Day Salary',
               f"No.of Days Present in {rep['month_en']} {year}", 'Total',
               'HR.', 'Salary /HR', '%', 'TOTAL',
               'HR.', 'Salary /HR', '%', 'TOTAL',
               'HR.', 'Salary /HR', '%', 'TOTAL',
               'TOTAL', 'GRAND TOTAL']
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = center
        cell.border = border
    header_row = r

    r += 1
    for row in rep['rows']:
        vals = [row['item'], row['sl_no'], row['name'], row['designation'],
                row['monthly_salary'], round(row['daily_salary'], 4),
                row['days_present'], row['basic_total'],
                row['ot_normal_hours'], round(row['ot_normal_rate'], 4),
                row['ot_normal_pct'], row['ot_normal_total'],
                row['ot_friday_hours'], round(row['ot_friday_rate'], 4),
                row['ot_friday_pct'], row['ot_friday_total'],
                row['ot_holiday_hours'], round(row['ot_holiday_rate'], 4),
                row['ot_holiday_pct'], row['ot_holiday_total'],
                row['ot_total'], row['grand_total']]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = border
            if i >= 5:
                cell.number_format = '0.000'
        r += 1

    t = rep['totals']
    ws.cell(row=r, column=19, value='GRAND TOTAL').font = Font(bold=True)
    gt = ws.cell(row=r, column=22, value=t['grand_total'])
    gt.font = Font(bold=True)
    gt.number_format = '0.000'

    widths = [6, 9, 30, 32, 11, 11, 13, 11] + [9, 10, 6, 11] * 3 + [11, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"manpower_{(c.get('company_name') or 'report')}_{rep['month_en']}_{year}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---------- إدارة العقود ----------

@manpower_bp.route('/api/manpower/contracts', methods=['GET'])
@login_required
@require_permission('report.manpower')
def contracts_api():
    return jsonify({'success': True,
                    'contracts': list_contracts(get_db_connection(), active_only=False)})


@manpower_bp.route('/api/manpower/contracts/save', methods=['POST'])
@login_required
@require_permission('report.manpower')
def save_contract():
    data = request.get_json(silent=True) or request.form
    cid = data.get('id')
    fields = ('company_name', 'ministry_line1', 'station_name', 'contract_no',
              'work_description', 'payment_ref')
    vals = [(data.get(f) or '').strip() for f in fields]
    if not vals[0]:
        return jsonify({'success': False, 'message': 'اسم الشركة مطلوب'}), 400

    def num(key, default):
        try:
            return float(data.get(key) or default)
        except (TypeError, ValueError):
            return default

    src = (data.get('data_source') or 'raw').lower()
    if src not in ('raw', 'approved'):
        src = 'raw'
    extra = [num('ot_normal', 1.25), num('ot_friday', 1.5),
             num('ot_holiday', 2.0), int(num('days_basis', 26)),
             int(num('hours_per_day', 8)), src]

    conn = get_db_connection()
    if cid:
        conn.execute(
            """UPDATE manpower_contracts SET company_name=?, ministry_line1=?,
               station_name=?, contract_no=?, work_description=?, payment_ref=?,
               ot_normal=?, ot_friday=?, ot_holiday=?, days_basis=?,
               hours_per_day=?, data_source=? WHERE id=?""", vals + extra + [cid])
    else:
        conn.execute(
            """INSERT INTO manpower_contracts
               (company_name, ministry_line1, station_name, contract_no,
                work_description, payment_ref, ot_normal, ot_friday,
                ot_holiday, days_basis, hours_per_day, data_source)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""", vals + extra)
    conn.commit()
    return jsonify({'success': True})


@manpower_bp.route('/api/manpower/contracts/<int:cid>/logo', methods=['POST'])
@login_required
@require_permission('report.manpower')
def upload_logo(cid):
    """رفع شعار الشركة.

    الامتداد يُتحقّق منه لا اسم الملف: ملفٌ يُرفع باسم آمن وامتداد تنفيذي
    داخل مجلد يخدمه الخادم ثغرةٌ لا خطأ تسمية. والاسم يُشتقّ من رقم العقد
    لا من اسم الملف المرفوع، فلا يُكتب خارج المجلد بمسار ملفّق.
    """
    f = request.files.get('logo')
    if not f or not f.filename:
        return jsonify({'success': False, 'message': 'لم يُرفع ملف'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_LOGO:
        return jsonify({'success': False,
                        'message': 'صيغة غير مسموحة: ' + ', '.join(sorted(ALLOWED_LOGO))}), 400

    folder = os.path.join('static', 'img', 'logos')
    os.makedirs(folder, exist_ok=True)
    name = f'contract_{cid}{ext}'
    f.save(os.path.join(folder, name))

    conn = get_db_connection()
    conn.execute('UPDATE manpower_contracts SET company_logo = ? WHERE id = ?',
                 (name, cid))
    conn.commit()
    return jsonify({'success': True, 'logo': name})
